import pandas as pd
import json

from db.db_helper import get_line_item_details


def create_lineitem_report(dbstring, doc_ids):

    detailed_rows = []

    summary_rows = []

    for doc_id in doc_ids:

        print(f"Processing LineItems: {doc_id}")

        headers = [f"Line Item Group {i}" for i in range(1, 50)]

        green_count = 0
        red_count = 0

        for header in headers:

            rows = get_line_item_details(
                doc_id,
                header,
                dbstring
            )

            if not rows:
                continue

            for item in rows:

                topic = item.get("topic", "")

                # Prefer edited value first
                value = item.get("topc_value_edited")

                if value is None or str(value).strip() == "":

                    value = item.get("topic_value", "")

                value = "" if value is None else str(value).strip()

                # Read meta_data
                meta_data = item.get("meta_data", b"{}")

                try:

                    # bytes -> string
                    if isinstance(meta_data, bytes):

                        meta_data = meta_data.decode("utf-8")

                    meta_json = json.loads(meta_data)

                    flag = meta_json.get("flag", "").lower()

                except Exception:

                    flag = ""

                # GREEN
                if flag == "g":

                    status = "GREEN"

                    green_count += 1

                # RED
                elif flag == "r":

                    status = "RED"

                    red_count += 1

                # UNKNOWN
                else:

                    status = "UNKNOWN"

                detailed_rows.append({

                    "doc_id": doc_id,
                    "header": header,
                    "topic": topic,
                    "value": value,
                    "flag": flag,
                    "status": status
                })

        total_fields = green_count + red_count

        accuracy = (
            (green_count / total_fields) * 100
            if total_fields
            else 0
        )

        summary_rows.append({

            "doc_id": doc_id,
            "total_fields": total_fields,
            "green_count": green_count,
            "red_count": red_count,
            "accuracy": round(accuracy, 2)
        })

    # Create DataFrames
    detailed_df = pd.DataFrame(detailed_rows)

    summary_df = pd.DataFrame(summary_rows)

    # AVG row
    avg_accuracy = summary_df["accuracy"].mean()

    avg_row = pd.DataFrame([{

        "doc_id": "AVG",
        "total_fields": "",
        "green_count": "",
        "red_count": "",
        "accuracy": round(avg_accuracy, 2)
    }])

    summary_df = pd.concat(
        [summary_df, avg_row],
        ignore_index=True
    )

    # Export Excel
    output_file = "reports/LineItems_report.xlsx"

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        detailed_df.to_excel(
            writer,
            sheet_name="Detailed_Report",
            index=False
        )

        summary_df.to_excel(
            writer,
            sheet_name="Summary_Report",
            index=False
        )

        # Excel Color Formatting
        worksheet = writer.sheets["Detailed_Report"]

        from openpyxl.styles import PatternFill

        green_fill = PatternFill(
            start_color="00FF00",
            end_color="00FF00",
            fill_type="solid"
        )

        red_fill = PatternFill(
            start_color="FF0000",
            end_color="FF0000",
            fill_type="solid"
        )

        gray_fill = PatternFill(
            start_color="C0C0C0",
            end_color="C0C0C0",
            fill_type="solid"
        )

        # Status column = F
        for row in range(2, worksheet.max_row + 1):

            cell = worksheet[f"F{row}"]

            if cell.value == "GREEN":

                cell.fill = green_fill

            elif cell.value == "RED":

                cell.fill = red_fill

            elif cell.value == "UNKNOWN":

                cell.fill = gray_fill

    print(f"\n✅ LineItem Report Generated: {output_file}")